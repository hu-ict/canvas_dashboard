import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

from scripts.lib.file import read_dashboard

columns = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

def create_picklist(list_items):
    options = []
    for list_item in list_items:
        options.append(list_item)
    # Voeg een picklist toe met opties
    # 4. Maak DataValidation objecten
    picklist = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    # Optioneel: zet een prompt en foutmelding
    picklist.promptTitle = "Selecteer een optie"
    picklist.prompt = "Kies uit de lijst"
    picklist.errorTitle = "Ongeldige invoer"
    picklist.error = "Selecteer een waarde uit de lijst"
    return picklist


def get_data_frame(assignment_group_names, student_groups):
# Bouw de data voor Excel
    rows = []
    print("LIBT25 -", assignment_group_names)
    print("LIBT26 -", student_groups)
    for student_group in student_groups:
        row = {}
        for assignment_group_name in assignment_group_names:
            if assignment_group_name == "Group":
                row[assignment_group_name] = student_group.name
            else:
                row[assignment_group_name] = ""
        rows.append(row)

    return pd.DataFrame(rows, columns=assignment_group_names)


def create_tab(ws, data_frame, teachers):
    print("LIBT30 -", data_frame)
    # print("LIBT31 -", teachers)
    for r in dataframe_to_rows(data_frame, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="left")
    picklist = create_picklist(teachers)
    # Stel kolombreedtes in
    ws.column_dimensions['A'].width = 75  # Kolom A breder
    # Voeg de validatie toe aan een cel (bijv. A1)
    ws.add_data_validation(picklist)
    column = columns[len(data_frame.columns)-1]
    row = len(data_frame)+1
    print("LIBT31 -", column, row)
    picklist.add(f"B2:{column}{row}")  # assignment1
    for column in columns[1:len(data_frame.columns)]:
        ws.column_dimensions[column].width = 15  # Kolom breder


def generate_trm(course_instance, config):
    print("GCF01 - generate_trm.py")

    teachers = []
    for teacher in config.teachers:
        teachers.append(teacher.name)

    # Maak kolomnamen (eerste kolom = 'group', daarna assignment group names)
    assignment_groups_1_names = ["Group"]
    for assignment_group in config.assignment_groups:
        if assignment_group.groups == "groups_1":
            assignment_groups_1_names.append(assignment_group.name)

    assignment_groups_2_names = ["Group"]
    for assignment_group in config.assignment_groups:
        if assignment_group.groups == "groups_2":
            assignment_groups_2_names.append(assignment_group.name)
    # Maak workbook en sheet
    wb = Workbook()
    # wb.create_sheet(title="Projects", index=0) # Vooraan
    ws = wb["Sheet"]
    ws.title = "groups_1"
    df_groups_1 = get_data_frame(assignment_groups_1_names, config.groups_1)
    create_tab(ws, df_groups_1, teachers)

    dashboard = read_dashboard(course_instance.get_dashboard_file_name())
    if len(dashboard.groups_2_name) > 0:
        wb.create_sheet(title="groups_2", index=1)  # Vooraan
        ws = wb["groups_2"]
        df_groups_2 = get_data_frame(assignment_groups_2_names, config.groups_2)
        create_tab(ws, df_groups_2, teachers)

    # wb.create_sheet(title="learning_outcomes", index=2)
    # ws = wb["learning_outcomes"]
    # ws["A1"] = "Id"
    # ws["B1"] = "Short"
    # ws["C1"] = "Description"
    # ws["A2"] = "LU1"
    # ws["A3"] = "LU2"
    # ws["A4"] = "LU3"
    # ws["A5"] = "LU4"
    # ws["A6"] = "LU5"

    # wb.create_sheet(title="perspectives", index=3)
    # ws = wb["perspectives"]
    # ws["A1"] = "name"
    # ws["A2"] = "level_moments"
    # ws["A3"] = "grade_moments"
    # ws["B1"] = "title"
    # ws["B2"] = "Peilmomenten"
    # ws["B3"] = "Beoordelingsmomenten"
    # ws["C1"] = "assignment_group_names"

    # wb.create_sheet(title="roles", index=4)
    # ws = wb["roles"]
    # ws["A1"] = "short"
    # ws["A2"] = "AI"
    # ws["A3"] = "BIM"
    # ws["A4"] = "CSC"
    # ws["A5"] = "SD"
    # ws["A6"] = "TI"
    # ws["B1"] = "name"
    # ws["B2"] = "Artificial Intelligence"
    # ws["B3"] = "Business and IT Management"
    # ws["B4"] = "Cyber Security and Cloud"
    # ws["B5"] = "Software Development"
    # ws["B6"] = "Technische Informatica"
    # ws["C1"] = "btn_color"
    # ws["C2"] = "border-warning"
    # ws["C3"] = "border-success"
    # ws["C4"] = "border-danger"
    # ws["C5"] = "border-dark"
    # ws["C6"] = "border-primary"

    # Sla het bestand op
    wb.save(course_instance.get_trm_file_name())
    print(f"LTRM11 - Excel-bestand '{course_instance.get_trm_file_name()}' is succesvol aangemaakt!")

