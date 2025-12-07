import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

from scripts.lib.file import read_course_instances
from lib.lib_date import get_actual_date

columns = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

def create_picklist(list_items):
    options = []
    for list_item in list_items:
        options.append(list_item)
    # Voeg een picklist toe met opties
    # 4. Maak DataValidation objecten
    picklist = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    # Optioneel: zet een prompt en foutmelding
    picklist.promptTitle = "Selecteer een optie"
    picklist.prompt = "Kies uit de lijst"
    picklist.errorTitle = "Ongeldige invoer"
    picklist.error = "Selecteer een waarde uit de lijst"
    return picklist


def get_data_frame(assignment_group_names, student_groups):
# Bouw de data voor Excel
    rows = []
    for student_group in student_groups:
        row = {}
        for assignment_group_name in assignment_group_names:
            if assignment_group_name == "Group":
                row[assignment_group_name] = student_group.name
            else:
                row[assignment_group_name] = ""
        rows.append(row)

    return pd.DataFrame(rows, columns=assignment_group_names)


def create_tab(ws, data_frame, teachers):
    for r in dataframe_to_rows(data_frame, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="left")
    picklist = create_picklist(teachers)
    # Stel kolombreedtes in
    ws.column_dimensions['A'].width = 75  # Kolom A breder
    # Voeg de validatie toe aan een cel (bijv. A1)
    ws.add_data_validation(picklist)
    column = columns[len(data_frame.columns)-1]
    row = len(data_frame)+1
    picklist.add(f"B2:{column}{row}")  # assignment1
    for column in columns[1:len(data_frame.columns)]:
        ws.column_dimensions[column].width = 15  # Kolom breder


def generate_course_xlsx(instance_name):
    print("GCX01 - generate_course_xlsx.py")
    g_actual_date = get_actual_date()
    instances = read_course_instances()
    excel_file = "course_xlsx.xlsx"
    # Maak workbook en sheet
    wb = Workbook()
    # wb.create_sheet(title="Projects", index=0)  # Vooraan
    ws = wb["Sheet"]
    ws.title = "learning_outcomes"
    ws["A1"] = "Id"
    ws["B1"] = "Short"
    ws["C1"] = "Description"
    ws["A2"] = "LU1"
    ws["A3"] = "LU2"
    ws["A4"] = "LU3"
    ws["A5"] = "LU4"
    ws["A6"] = "LU5"

    wb.create_sheet(title="perspectives", index=1)
    ws = wb["perspectives"]
    ws["A1"] = "name"
    ws["A2"] = "level_moments"
    ws["A3"] = "grade_moments"
    ws["B1"] = "title"
    ws["B2"] = "Peilmomenten"
    ws["B3"] = "Beoordelingsmomenten"
    ws["C1"] = "assignment_group_names"

    wb.create_sheet(title="roles", index=1)
    ws = wb["roles"]
    ws["A1"] = "short"
    ws["A2"] = "AI"
    ws["A3"] = "BIM"
    ws["A4"] = "CSC"
    ws["A5"] = "SD"
    ws["A6"] = "TI"
    ws["B1"] = "name"
    ws["B2"] = "Artificial Intelligence"
    ws["B3"] = "Business and IT Management"
    ws["B4"] = "Cyber Security and Cloud"
    ws["B5"] = "Software Development"
    ws["B6"] = "Technische Informatica"
    ws["C1"] = "btn_color"
    ws["C2"] = "border-warning"
    ws["C3"] = "border-success"
    ws["C4"] = "border-danger"
    ws["C5"] = "border-dark"
    ws["C6"] = "border-primary"
    wb.save(excel_file)
    print(f"Excel-bestand '{excel_file}' is succesvol aangemaakt!")

    print("GCX99 Time running:", (get_actual_date() - g_actual_date).seconds, "seconds")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        generate_course_xlsx(sys.argv[1])
    else:
        generate_course_xlsx("")