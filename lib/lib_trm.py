import json
import sys
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from lib.file import read_environment, read_secret_api_key, read_course
from lib.lib_date import get_actual_date, API_URL
from model.environment.Environment import ENVIRONMENT_FILE_NAME

columns = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def create_picklist(list_items):
    options = []
    for list_item in list_items:
        options.append(list_item)
    # Voeg een picklist toe met opties
    # 4. Maak DataValidation objecten
    picklist = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    # Optioneel: zet een prompt en foutmelding
    picklist.promptTitle = "Selecteer een optie"
    picklist.prompt = "Kies uit de lijst"
    picklist.errorTitle = "Ongeldige invoer"
    picklist.error = "Selecteer een waarde uit de lijst"
    return picklist


def get_data_frame(assignment_group_names, student_groups):
# Bouw de data voor Excel
    rows = []
    for student_group in student_groups:
        row = {}
        for assignment_group_name in assignment_group_names:
            if assignment_group_name == "Group":
                row[assignment_group_name] = student_group.name
            else:
                row[assignment_group_name] = ""
        rows.append(row)

    return pd.DataFrame(rows, columns=assignment_group_names)


def create_tab(ws, data_frame, teachers):
    # print("GTRM11 -", teachers)
    for r in dataframe_to_rows(data_frame, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="left")
    picklist = create_picklist(teachers)
    # Stel kolombreedtes in
    ws.column_dimensions['A'].width = 75  # Kolom A breder
    # Voeg de validatie toe aan een cel (bijv. A1)
    ws.add_data_validation(picklist)
    column = columns[len(data_frame.columns)-1]
    row = len(data_frame)+1
    picklist.add(f"B2:{column}{row}")  # assignment1
    for column in columns[1:len(data_frame.columns)]:
        ws.column_dimensions[column].width = 15  # Kolom breder


def generate_trm(course_instance, config):
    print("GCF01 - generate_trm.py")

    teachers = []
    for teacher in config.teachers:
        teachers.append(teacher.name)

    # Maak kolomnamen (eerste kolom = 'group', daarna assignment group names)
    assignment_group_names = ["Group"]
    for assignment_group in config.assignment_groups:
        assignment_group_names.append(assignment_group.name)

    # Maak workbook en sheet
    wb = Workbook()
    # wb.create_sheet(title="Projects", index=0)  # Vooraan
    ws = wb["Sheet"]
    ws.title = "Projects"
    df_projects = get_data_frame(assignment_group_names, config.project_groups)

    create_tab(ws, df_projects, teachers)

    wb.create_sheet(title="Guilds", index=1)  # Vooraan
    ws = wb["Guilds"]
    df_guilds = get_data_frame(assignment_group_names, config.guild_groups)
    create_tab(ws, df_guilds, teachers)

    # Sla het bestand op
    wb.save(course_instance.get_trm_file_name())
    print(f"LTRM11 - Excel-bestand '{course_instance.get_trm_file_name()}' is succesvol aangemaakt!")
