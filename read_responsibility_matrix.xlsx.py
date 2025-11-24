import json
import sys

from canvasapi import Canvas

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

from generate_students import get_groups
from lib.file import read_course_instances, read_start, read_config_from_canvas, read_course
from lib.lib_date import get_actual_date, API_URL
from model.AssignmentGroup import AssignmentGroup
from model.teacher.AssignmentGroupNameId import AssignmentGroupNameId
from model.teacher.Responsibility import Responsibility
from model.teacher.ResponsibilityMatrix import ResponsibilityMatrix
from model.teacher.Teacher import Teacher

columns = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

def read_responsibility_matrix(instance_name):
    print("RRM01 - read_responsibility_matrix.py")
    g_actual_date = get_actual_date()
    instances = read_course_instances()
    if len(instance_name) > 0:
        instances.current_instance = instance_name
    instance = instances.get_instance_by_name(instances.current_instance)
    print("Instance:", instance.name)
    start = read_start(instance.get_start_file_name())
    # Initialize a new Canvas object
    canvas = Canvas(API_URL, start.api_key)
    user = canvas.get_current_user()
    print(user.name)
    canvas = Canvas(API_URL, start.api_key)
    canvas_course = canvas.get_course(start.canvas_course_id)
    if start.config_target == "CANVAS":
        course = read_config_from_canvas(canvas_course)
    else:
        course = read_course(instance.get_config_file_name())

    excel_file = instance.get_project_path()+"responsibility_matrix.xlsx"
    # Open het Excel-bestand
    wb = load_workbook(excel_file)
    ws = wb["Projects"]
    assignment_groups = {}
    for header in ws[1][1:]:
        assignment_groups[header.column] = header.value
        print("RRM11 -", header.column, header.value)
    for row in ws.iter_rows(min_row=2):
        group_name = ""
        for cell in row:
            if cell.column in assignment_groups:
                print("RRM12 -", group_name, assignment_groups[cell.column], cell.value)
                assignment_group = course.find_assignment_group_by_name(assignment_groups[cell.column])
                group = course.find_project_group_by_name(group_name)
                teacher = course.find_teacher_by_name(cell.value)
                if teacher:
                    print("RRM13 -", assignment_group)
                    print("RRM14 -", group, group_name)
                    print("RRM15 -", teacher)
                    responsibility = Responsibility("project_groups", [group.name], assignment_group.id)
                    teacher.put_responsibility(responsibility)
            else:
                group_name = cell.value
                print("RRM17 - group_name", group_name)

    ws = wb["Guilds"]
    assignment_groups = {}
    for header in ws[1][1:]:
        assignment_groups[header.column] = header.value
        print("RRM21 -", header.column, header.value)
    for row in ws.iter_rows(min_row=2):
        group_name = ""
        for cell in row:
            if cell.column in assignment_groups:
                print("RRM22 -", group_name, assignment_groups[cell.column], cell.value)
                assignment_group = course.find_assignment_group_by_name(assignment_groups[cell.column])
                group = course.find_guild_group_by_name(group_name)
                for group in course.guild_groups:
                    print("RRM28 - group_name", group.name)
                teacher = course.find_teacher_by_name(cell.value)
                if teacher:
                    print("RRM23 -", assignment_group)
                    print("RRM24 -", group, group_name)
                    print("RRM25 -", teacher)
                    responsibility = Responsibility("guild_groups", [group.name], assignment_group.id)
                    teacher.put_responsibility(responsibility)
            else:
                group_name = cell.value
                print("RRM27 - group_name", group_name)

    print(f"Excel-bestand '{excel_file}' is succesvol gelezen!")
    print("RRM98 - ConfigFileName:", instance.get_config_file_name())
    with open(instance.get_config_file_name(), 'w') as f:
        dict_result = course.to_json()
        json.dump(dict_result, f, indent=2)

    print("RRM99 Time running:", (get_actual_date() - g_actual_date).seconds, "seconds")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        read_responsibility_matrix(sys.argv[1])
    else:
        read_responsibility_matrix("")