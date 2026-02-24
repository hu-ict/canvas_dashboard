import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from scripts.lib.file import read_course, read_environment
from scripts.lib.file_const import ENVIRONMENT_FILE_NAME
from scripts.lib.lib_date import get_actual_date

columns = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

def clean_up_responsibilities(teachers):
    for teacher in teachers:
        teacher.responsibilities = []


def read_trm(course_code, instance_name):
    print("GCF01 - generate_config.py")
    g_actual_date = get_actual_date()
    print(Path(__file__).resolve().parent)
    BASE_DIR = Path(__file__).resolve().parent
    environment = read_environment(ENVIRONMENT_FILE_NAME)
    current_instance = environment.get_instance_of_course(environment.current_instance)
    execution = environment.get_execution_by_name("env_1")

    course = read_course(current_instance.get_config_file_name())
    clean_up_responsibilities(course.teachers)

    # Open het Excel-bestand
    wb = load_workbook(current_instance.get_trm_file_name())
    ws = wb["projects"]
    assignment_groups = {}
    for header in ws[1][1:]:
        assignment_groups[header.column] = header.value
        print("RRM11 -", header.column, header.value)
    for row in ws.iter_rows(min_row=2):
        student_group_name = ""
        for cell in row:
            if cell.column in assignment_groups:
                print("RRM12 -", student_group_name, assignment_groups[cell.column], cell.value)
                assignment_group = course.find_assignment_group_by_name(assignment_groups[cell.column])
                # group = course.find_project_group_by_name(student_group_name)
                teacher = course.find_teacher_by_name(cell.value)
                if teacher:
                    print("RRM13 -", assignment_group)
                    print("RRM14 -", student_group_name)
                    print("RRM15 -", teacher)
                    teacher.put_responsibility("project_groups", student_group_name, assignment_group.id)
            else:
                student_group_name = cell.value
                print("RRM17 - student_group_name", student_group_name)

    ws = wb["guilds"]
    assignment_groups = {}
    for header in ws[1][1:]:
        assignment_groups[header.column] = header.value
        print("RRM21 -", header.column, header.value)
    for row in ws.iter_rows(min_row=2):
        student_group_name = ""
        for cell in row:
            if cell.column in assignment_groups:
                print("RRM22 -", student_group_name, assignment_groups[cell.column], cell.value)
                assignment_group = course.find_assignment_group_by_name(assignment_groups[cell.column])
                # group = course.find_guild_group_by_name(student_group_name)
                teacher = course.find_teacher_by_name(cell.value)
                if teacher:
                    # print("RRM23 -", assignment_group)
                    # print("RRM24 -", student_group_name)
                    # print("RRM25 -", teacher)
                    teacher.put_responsibility("guild_groups", student_group_name, assignment_group.id)
            else:
                student_group_name = cell.value
                print("RRM27 - group_name", student_group_name)

    # opschonen van teachers zonder "responsibilities"
    for teacher in course.teachers[:]:  # kopie met slicing
        if len(teacher.responsibilities) == 0:
            print("RRM31 - Verwijder teacher", teacher.name)
            course.teachers.remove(teacher)
    print(f"Excel-bestand '{current_instance.get_trm_file_name()}' is succesvol gelezen!")
    print("RRM98 - ConfigFileName:", current_instance.get_config_file_name())
    with open(current_instance.get_config_file_name(), 'w') as f:
        dict_result = course.to_json()
        json.dump(dict_result, f, indent=2)

    print("RRM99 Time running:", (get_actual_date() - g_actual_date).seconds, "seconds")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        read_trm(sys.argv[1], sys.argv[1])
    else:
        read_trm("", "")