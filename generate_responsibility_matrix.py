import json
import sys

from canvasapi import Canvas

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

from generate_students import get_groups
from lib.file import read_course_instances, read_start
from lib.lib_date import get_actual_date, API_URL
from model.AssignmentGroup import AssignmentGroup
from model.teacher.AssignmentGroupNameId import AssignmentGroupNameId
from model.teacher.ResponsibilityMatrix import ResponsibilityMatrix
from model.teacher.Teacher import Teacher

columns = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

def create_picklist(list_items):
    options = []
    for list_item in list_items:
        options.append(list_item)
    # Voeg een picklist toe met opties
    # 4. Maak DataValidation objecten
    picklist = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    # Optioneel: zet een prompt en foutmelding
    picklist.promptTitle = "Selecteer een optie"
    picklist.prompt = "Kies uit de lijst"
    picklist.errorTitle = "Ongeldige invoer"
    picklist.error = "Selecteer een waarde uit de lijst"
    return picklist


def get_data_frame(assignment_group_names, student_groups):
# Bouw de data voor Excel
    rows = []
    for student_group in student_groups:
        row = {}
        for assignment_group_name in assignment_group_names:
            if assignment_group_name == "Group":
                row[assignment_group_name] = student_group.name
            else:
                row[assignment_group_name] = ""
        rows.append(row)

    return pd.DataFrame(rows, columns=assignment_group_names)


def create_tab(ws, data_frame, teachers):
    for r in dataframe_to_rows(data_frame, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="left")
    picklist = create_picklist(teachers)
    # Stel kolombreedtes in
    ws.column_dimensions['A'].width = 75  # Kolom A breder
    # Voeg de validatie toe aan een cel (bijv. A1)
    ws.add_data_validation(picklist)
    column = columns[len(data_frame.columns)-1]
    row = len(data_frame)+1
    picklist.add(f"B2:{column}{row}")  # assignment1
    for column in columns[1:len(data_frame.columns)]:
        ws.column_dimensions[column].width = 15  # Kolom breder


def generate_responsibility_matrix(instance_name):
    print("GRM01 - generate_responsibility_matrix.py")
    g_actual_date = get_actual_date()
    instances = read_course_instances()
    if len(instance_name) > 0:
        instances.current_instance = instance_name
    instance = instances.get_instance_by_name(instances.current_instance)
    print("Instance:", instance.name)
    start = read_start(instance.get_start_file_name())
    # Initialize a new Canvas object
    canvas = Canvas(API_URL, start.api_key)
    user = canvas.get_current_user()
    print(user.name)
    canvas_course = canvas.get_course(start.canvas_course_id)
    responsibility_matrix_guild = ResponsibilityMatrix(start.guild_group_name)
    responsibility_matrix_project = ResponsibilityMatrix(start.project_group_name)
    canvas_assignment_groups = canvas_course.get_assignment_groups(include=['assignments', 'overrides'])
    for canvas_assignment_group in canvas_assignment_groups:
        assignment_group = AssignmentGroupNameId(canvas_assignment_group.id, canvas_assignment_group.name, "project")
        responsibility_matrix_project.assignment_groups.append(assignment_group)
        print("GRM05 - assignment_group", assignment_group)

    # retrieve Teachers
    canvas_users = canvas_course.get_users(enrollment_type=['teacher', 'ta'])
    teacher_count = 0
    teachers = []
    for canvas_user in canvas_users:
        teacher_count += 1
        if hasattr(canvas_user, 'email'):
            teacher = Teacher(canvas_user.id, canvas_user.name, canvas_user.email)
            # print("GCONF15 - Create teacher with email", canvas_user.email)
        else:
            teacher = Teacher(canvas_user.id, canvas_user.name, "")
            print("GRM07 - Create teacher without email", canvas_user.name)
        responsibility_matrix_project.teachers.append(teacher)
        teachers.append(teacher.name)
        print("GRM09 - Teacher", teacher)

    groups = get_groups(start.project_group_name, canvas_course)
    responsibility_matrix_project.student_groups = groups

    # Maak kolomnamen (eerste kolom = 'group', daarna assignment group names)
    assignment_group_names = ["Group"]
    for assignment_group in responsibility_matrix_project.assignment_groups:
        assignment_group_names.append(assignment_group.name)

    project_groups = get_groups(start.project_group_name, canvas_course)
    responsibility_matrix_project.student_groups = project_groups
    guild_groups = get_groups(start.guild_group_name, canvas_course)
    responsibility_matrix_guild.student_groups = guild_groups

    # Schrijf naar Excel

    excel_file = instance.get_project_path()+"responsibility_matrix.xlsx"
    # Maak workbook en sheet
    wb = Workbook()
    # wb.create_sheet(title="Projects", index=0)  # Vooraan
    ws = wb["Sheet"]
    ws.title = "Projects"
    df_projects = get_data_frame(assignment_group_names, project_groups)
    create_tab(ws, df_projects, teachers)

    wb.create_sheet(title="Guilds", index=1)  # Vooraan
    ws = wb["Guilds"]
    df_guilds = get_data_frame(assignment_group_names, guild_groups)
    create_tab(ws, df_guilds, teachers)
    # wb.create_sheet(title="Perspectives", index=0)  # Vooraan
    # ws = wb["Perspectives"]
    # ws.append(["portfolio"])
    # ws.append(["level_moments"])
    # ws.append(["grade_moments"])
    # ws.column_dimensions['A'].width = 30  # Kolom A breder
    #
    # picklist = create_picklist(assignment_group_names[1:])
    # # Voeg de validatie toe aan een cel (bijv. A1)
    # ws.add_data_validation(picklist)
    # column = "H"
    # row = 3
    # picklist.add(f"B1:{column}{row}")  # assignment1
    # # Sla het bestand op
    wb.save(excel_file)
    print(f"Excel-bestand '{excel_file}' is succesvol aangemaakt!")

    print("GRM99 Time running:", (get_actual_date() - g_actual_date).seconds, "seconds")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        generate_responsibility_matrix(sys.argv[1])
    else:
        generate_responsibility_matrix("")