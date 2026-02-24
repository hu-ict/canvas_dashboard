import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

from scripts.lib.file import read_environment
from scripts.lib.file_const import ENVIRONMENT_FILE_NAME
from scripts.lib.lib_date import get_actual_date
from scripts.lib.lib_trm import create_picklist, columns


def generate_course_xlsx(instance_name):
    print("GCX01 - generate_course_xlsx.py")
    environment = read_environment(ENVIRONMENT_FILE_NAME)
    current_instance = environment.get_instance_of_course(environment.current_instance)
    execution = environment.get_execution_by_name("env_1")

    g_actual_date = get_actual_date()
    # Maak workbook en sheet
    wb = Workbook()
    wb.create_sheet(title="Projects", index=0)  # Vooraan
    ws = wb["Sheet"]
    ws.title = "learning_outcomes"
    ws["A1"] = "Id"
    ws["B1"] = "Short"
    ws["C1"] = "Description"
    ws["A2"] = "LU1"
    ws["A3"] = "LU2"
    ws["A4"] = "LU3"
    ws["A5"] = "LU4"
    ws["A6"] = "LU5"

    wb.create_sheet(title="perspectives", index=1)
    ws = wb["perspectives"]
    ws["A1"] = "name"
    ws["A2"] = "level_moments"
    ws["A3"] = "grade_moments"
    ws["B1"] = "title"
    ws["B2"] = "Peilmomenten"
    ws["B3"] = "Beoordelingsmomenten"
    ws["C1"] = "assignment_group_names"

    wb.create_sheet(title="roles", index=1)
    ws = wb["roles"]
    ws["A1"] = "short"
    ws["A2"] = "AI"
    ws["A3"] = "BIM"
    ws["A4"] = "CSC"
    ws["A5"] = "SD"
    ws["A6"] = "TI"
    ws["B1"] = "name"
    ws["B2"] = "Artificial Intelligence"
    ws["B3"] = "Business and IT Management"
    ws["B4"] = "Cyber Security and Cloud"
    ws["B5"] = "Software Development"
    ws["B6"] = "Technische Informatica"
    ws["C1"] = "btn_color"
    ws["C2"] = "border-warning"
    ws["C3"] = "border-success"
    ws["C4"] = "border-danger"
    ws["C5"] = "border-dark"
    ws["C6"] = "border-primary"
    wb.save(current_instance.get_trm_file_name())
    print(f"Excel-bestand '{current_instance.get_trm_file_name()}' is succesvol aangemaakt!")

    print("GCX99 Time running:", (get_actual_date() - g_actual_date).seconds, "seconds")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        generate_course_xlsx(sys.argv[1])
    else:
        generate_course_xlsx("")